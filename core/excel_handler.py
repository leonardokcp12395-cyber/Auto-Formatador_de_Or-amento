from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
from copy import copy
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.page import PageMargins
from decimal import Decimal, InvalidOperation, ROUND_CEILING, ROUND_FLOOR, ROUND_HALF_UP
import gc
import os
import shutil
import re
import math
from typing import Optional, Dict, List, Tuple, Any, Callable
from utils.logger import Logger
from core.paths import get_app_dir
from core.exceptions import ExcelProcessError, DataExtractionError, TemplateNotFoundError



class OrcamentoEngine:
    def __init__(self, config: Dict[str, Any] = None):
        self.output_dir: str = str(get_app_dir() / "Output")
        if not os.path.exists(self.output_dir): 
            os.makedirs(self.output_dir)
        
        self.wb_out: Any = None
        self.ws_out: Any = None
        self.wb_src: Any = None
        self.ws_src: Any = None
        
        self.info: Dict[str, Any] = {}
        self.mapa_colunas: Dict[str, str] = {}
        self.perfil: Dict[str, Any] = {}
        self.mapa_saida: Dict[str, int] = {}
        self._item_total_original_sum: Decimal = Decimal("0")
        self.FMT_CONTABIL: str = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"??_);_(@_)'

    def gerar_excel_final(
        self,
        linhas_aprovadas: List[Dict[str, Any]],
        modelo_path: Optional[str],
        mapa_colunas: Dict[str, str],
        info: Dict[str, Any],
        progress_callback: Optional[Callable[[int], None]] = None,
        perfil: Optional[Dict[str, Any]] = None,
    ) -> Tuple[bool, str, Dict[str, Any]]:
        Logger.info(">>> PLANIFY ENGINE V50: TYPED & SAFE <<<")
        self.info = info
        self.mapa_colunas = mapa_colunas
        self.perfil = perfil or {}
        self.mapa_saida = self._normalizar_mapa_saida(self.perfil.get("mapa_colunas", {}))
        self._item_total_original_sum = Decimal("0")

        effective_model_path = self.perfil.get("caminho_template") or modelo_path
        
        if not effective_model_path or not os.path.exists(effective_model_path):
            raise TemplateNotFoundError(f"Template '{effective_model_path}' não foi encontrado.")

        save_path = ""

        try:
            ok, save_path = self._preparar_arquivo(effective_model_path)
            if not ok: 
                return False, save_path, {}
            
            self._processar_cabecalho()

            start_row = self._obter_linha_inicio()
            current_row, mapa_linhas = self._processar_itens(linhas_aprovadas, start_row, progress_callback)
            self._inserir_formulas_totais(mapa_linhas)
            self._processar_rodape(current_row, start_row)
            
            try:
                if self.wb_out:
                    self._sanitizar_mesclagens()
                    self._configurar_pagina_para_pdf()
                    self.wb_out.save(save_path)
            except PermissionError as e:
                raise ExcelProcessError(f"O arquivo '{save_path}' está aberto em outro programa. Feche-o e tente novamente.", e)

            Logger.info(f"✅ Concluído: {save_path}")
            return True, save_path, {}

        except ExcelProcessError as e:
            Logger.error(f"Erro ExcelProcessError: {e}")
            return False, str(e), {}
        except Exception as e:
            Logger.error(f"Erro Inesperado Engine: {e}")
            import traceback
            traceback.print_exc()
            return False, f"Erro crítico: {str(e)}", {}
        finally:
            self._cleanup()

    def _normalizar_mapa_saida(self, mapa_colunas: Dict[str, Any]) -> Dict[str, int]:
        """
        Converte o mapa configurável do perfil (A, B, AA...) em índices OpenPyXL.
        Defaults mantêm compatibilidade com os templates Planify históricos.
        """
        default_map = {
            "ITEM": "A",
            "CODIGO": "B",
            "BANCO": "C",
            "DESCRICAO": "D",
            "UNID": "E",
            "QUANT": "F",
            "UNIT": "G",
            "TOTAL": "H",
        }

        normalized: Dict[str, int] = {}
        merged_map = mapa_colunas or default_map

        for key, value in merged_map.items():
            canonical_key = str(key).upper().strip()
            try:
                if isinstance(value, int):
                    column_index = value
                else:
                    column_index = column_index_from_string(str(value).strip().upper())

                if column_index > 0:
                    normalized[canonical_key] = column_index
            except Exception:
                Logger.warning(f"Coluna de saída inválida ignorada: {key}={value}")

        return normalized

    def _obter_col_saida(self, campo: str) -> Optional[int]:
        return self.mapa_saida.get(campo)

    def _obter_letra_col_saida(self, campo: str) -> Optional[str]:
        col = self._obter_col_saida(campo)
        return get_column_letter(col) if col else None

    def _obter_linha_inicio(self) -> int:
        try:
            start_row = int(self.perfil.get("linha_inicio", 0))
            if start_row > 0:
                return start_row
        except (TypeError, ValueError):
            pass

        return self._encontrar_inicio_tabela()
            
    def _cleanup(self) -> None:
        """
        Libera handles de arquivo do OpenPyXL de forma determinística.
        Não existe COM aqui, mas workbooks carregados mantêm ZIP handles abertos.
        """
        for workbook_attr in ("wb_src", "wb_out"):
            workbook = getattr(self, workbook_attr, None)
            if workbook:
                try:
                    workbook.close()
                except Exception as exc:
                    Logger.warning(f"Falha ao fechar workbook {workbook_attr}: {exc}")
            setattr(self, workbook_attr, None)

        self.ws_src = None
        self.ws_out = None
        gc.collect()

    def _sanitizar_mesclagens(self) -> None:
        if not self.ws_out:
            return

        occupied = set()
        for merged in list(self.ws_out.merged_cells.ranges):
            merge_ref = str(merged)
            try:
                min_col, min_row, max_col, max_row = range_boundaries(merge_ref)
            except Exception:
                try:
                    self.ws_out.unmerge_cells(merge_ref)
                except Exception:
                    pass
                continue

            if min_col > max_col or min_row > max_row or (min_col == max_col and min_row == max_row):
                try:
                    self.ws_out.unmerge_cells(merge_ref)
                except Exception:
                    pass
                continue

            cells = {
                (row, col)
                for row in range(min_row, max_row + 1)
                for col in range(min_col, max_col + 1)
            }
            if occupied.intersection(cells):
                try:
                    self.ws_out.unmerge_cells(merge_ref)
                except Exception:
                    pass
                continue

            occupied.update(cells)

    def _normalizar_caminho_saida(self, requested_path: Any, output_suffix: str) -> str:
        save_path = str(requested_path or "").strip()
        if not save_path:
            return ""

        path_obj = os.path.abspath(os.path.expanduser(save_path))
        base, ext = os.path.splitext(path_obj)
        ext_lower = ext.lower()

        if ext_lower not in {".xlsx", ".xlsm"}:
            path_obj = f"{path_obj}{output_suffix}" if not ext else f"{base}{output_suffix}"
        elif output_suffix == ".xlsm" and ext_lower != ".xlsm":
            path_obj = f"{base}.xlsm"

        os.makedirs(os.path.dirname(path_obj), exist_ok=True)
        return path_obj

    def _configurar_pagina_para_pdf(self) -> None:
        if not self.ws_out:
            return

        try:
            self.ws_out.sheet_properties.pageSetUpPr.fitToPage = True
            self.ws_out.page_setup.fitToWidth = 1
            self.ws_out.page_setup.fitToHeight = 0
            self.ws_out.page_setup.orientation = getattr(self.ws_out, "ORIENTATION_LANDSCAPE", "landscape")
            self.ws_out.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75)
            self.ws_out.print_options.horizontalCentered = True

            max_output_col = max(self.mapa_saida.values(), default=8)
            max_output_col = max(max_output_col, 8)
            self.ws_out.print_area = f"A1:{get_column_letter(max_output_col)}{self.ws_out.max_row}"
        except Exception as exc:
            Logger.warning(f"Falha ao aplicar configuração de página para PDF: {exc}")

    def _preparar_arquivo(self, modelo_path: str) -> Tuple[bool, str]:
        """Cria cópia do modelo. Retorna o sucesso e o path salvo."""
        nome_base = self.info.get('nome_arquivo', 'Orcamento').strip()
        nome_base = re.sub(r'[<>:"/\\|?*]', '', nome_base)
        output_suffix = ".xlsm" if str(modelo_path).lower().endswith(".xlsm") else ".xlsx"

        requested_save_path = self._normalizar_caminho_saida(self.info.get("save_path"), output_suffix)
        if requested_save_path:
            save_path = requested_save_path
            if os.path.abspath(save_path).lower() == os.path.abspath(modelo_path).lower():
                raise ExcelProcessError("O arquivo final não pode sobrescrever o template do perfil.")
        else:
            tentativas = 0
            save_path = ""

            while tentativas < 20:
                sufixo = "" if tentativas == 0 else f"_v{tentativas}"
                nome_final = f"{nome_base}{sufixo}{output_suffix}"
                save_path = os.path.join(self.output_dir, nome_final)

                if os.path.exists(save_path):
                    try:
                        with open(save_path, 'a+'): pass
                    except IOError:
                        Logger.warning(f"Arquivo '{nome_final}' está aberto. Tentando v{tentativas+1}...")
                        tentativas += 1
                        continue
                break

            if tentativas >= 20:
                raise ExcelProcessError(f"Muitos arquivos '{nome_base}' bloqueados! Feche o aplicativo Excel.")
        
        try:
            import openpyxl
            shutil.copy(modelo_path, save_path)
            keep_vba = output_suffix == ".xlsm"
            self.wb_out = openpyxl.load_workbook(save_path, keep_vba=keep_vba)
            self.ws_out = self.wb_out.active
            self.wb_src = openpyxl.load_workbook(modelo_path, keep_vba=keep_vba)
            self.ws_src = self.wb_src.active
            return True, save_path
        except Exception as e:
            raise ExcelProcessError(f"Falha ao carregar e copiar template: {e}")

    def _processar_cabecalho(self) -> None:
        if self.perfil and not self.perfil.get("usa_cabecalho_padrao", False):
            return

        campus = self._header_value('campus')
        setor = self._header_value('setor')
        servidor = self._header_value('servidor')
        elaborador = self._header_value('elaborador')
        estagiario = self._header_value('estagiario')
        descricao_header = self._header_value('descricao_header', upper=True)
        data = self._header_value('data')
        orcafascio = self._header_value('orcafascio', default="XXXXX")
        processo = self._header_value('processo')
        fiscal = self._header_value('fiscal')
        num_orcamento = self._header_value('num_orcamento', default="XX")
        data_emissao = self._header_value('data_emissao')
        data_inicio = self._header_value('data_inicio')
        prazo = self._header_value('prazo')
        empenho = self._header_value('empenho')

        self._write_cell('A8', f"CAMPUS: {campus}", bold=True)
        self._write_cell('A9', f"SETOR:  {setor}", bold=True)
        self._write_cell('A10', f"SERVIDOR: {servidor}", bold=True)
        self._write_cell('A13', f"ORÇAMENTO ELABORADO POR: {elaborador}", bold=True)
        self._write_cell('A14', f"ESTAGIÁRIO: {estagiario}", bold=True)
        self._write_cell('C15', descricao_header, bold=False)
        self._write_cell('A18', f"DATA DE ELABORAÇÃO DO ORÇAMENTO: {data} (VALIDADE: 45 DIAS)", bold=True)
        self._write_cell('E18', f"CÓDIGO ORÇAFASCIO:  {orcafascio}", bold=True)
        self._write_cell('E21', f"NÚMERO DO PROCESSO:  {processo}", bold=True)
        self._write_cell('D22', f"FISCAL DO SERVIÇO: {fiscal}", bold=True)
        
        self._write_cell('C20', f"Nº {num_orcamento}", bold=True)
        self._write_cell('D20', f"DATA DE EMISSÃO: {data_emissao}", bold=True)
        self._write_cell('D21', f"DATA DE INÍCIO: {data_inicio}", bold=True)
        self._write_cell('E20', f"PRAZO: {prazo}", bold=True)
        self._write_cell('G20', f"EMPENHO: {empenho}", bold=True)

    def _header_value(self, key: str, default: str = "XXXXX", upper: bool = False) -> str:
        value = self.info.get(key)
        if value is None:
            text = ""
        else:
            text = str(value).strip()

        if not text or text.lower() in {"none", "nan", "null"}:
            text = default

        return text.upper() if upper else text

    def _encontrar_inicio_tabela(self) -> int:
        start_row = 15
        if not self.ws_out: return start_row
        for r in range(1, 50):
            val = str(self.ws_out.cell(r, 4).value).upper()
            if 'DESCRIÇÃO' in val or 'DISCRIMINAÇÃO' in val:
                start_row = r + 1
                break
        return start_row

    def _processar_itens(self, linhas_aprovadas: List[Dict[str, Any]], start_row: int, progress_callback: Optional[Callable[[int], None]] = None) -> Tuple[int, List[Dict[str, Any]]]:
        FMT_NUM = '0.########'
        FMT_MOEDA = '"R$ "#,##0.00'
        calc_mode = self.info.get('calc_mode', 'EXACT')
        altura_base = self.info.get('altura_linha', 24.75)

        cols = {k: self.mapa_colunas.get(k, k) for k in ["ITEM","CODIGO","BANCO","DESCRICAO","UNID","QUANT","UNIT","TOTAL"]}
        mapa_linhas_escritas: List[Dict[str, Any]] = []
        current_row = start_row
        total_linhas = len(linhas_aprovadas)

        for i, row_data in enumerate(linhas_aprovadas):
            self._limpar_mesclagem_linha(current_row)
            nivel = row_data.get("_NIVEL_FORCADO", "ITEM")
            descricao = row_data.get(cols["DESCRICAO"], '')

            for campo in ["ITEM", "CODIGO", "BANCO", "DESCRICAO"]:
                output_col = self._obter_col_saida(campo)
                if output_col:
                    self._safe_write(current_row, output_col, row_data.get(cols[campo], ''))

            if nivel == "ITEM":
                unid_col = self._obter_col_saida("UNID")
                quant_col = self._obter_col_saida("QUANT")
                unit_col = self._obter_col_saida("UNIT")
                total_col = self._obter_col_saida("TOTAL")

                if unid_col:
                    self._safe_write(current_row, unid_col, row_data.get(cols["UNID"], ''))

                try:
                    qtd_raw = self._parse_num(row_data.get(cols["QUANT"]))
                    unit_raw = self._parse_num(row_data.get(cols["UNIT"]))
                    total_original = self._parse_num(row_data.get(cols["TOTAL"]))
                    qtd_final = self._aplicar_precisao(qtd_raw, calc_mode)
                    unit_final = self._aplicar_precisao(unit_raw, calc_mode)
                except Exception as e:
                    raise DataExtractionError(f"Erro de extração numérica na linha excel {current_row}: {e}")

                if qtd_final is not None and quant_col:
                    self._safe_write(current_row, quant_col, qtd_final, FMT_NUM)
                if unit_final is not None and unit_col:
                    self._safe_write(current_row, unit_col, unit_final, FMT_MOEDA)
                if total_col:
                    total_value = self._get_exact_formula(
                        row=current_row,
                        col_q=quant_col,
                        col_u=unit_col,
                        target_total=total_original,
                        q_val=qtd_final,
                        u_val=unit_final,
                    )
                    if total_value is not None:
                        self._safe_write(current_row, total_col, total_value, self.FMT_CONTABIL)
                    if total_original is not None:
                        self._item_total_original_sum += total_original

                mapa_linhas_escritas.append({'row': current_row, 'nivel': 'ITEM'})
            else:
                mapa_linhas_escritas.append({'row': current_row, 'nivel': nivel})

            self._aplicar_estilo_hierarquico(current_row, nivel)
            self._ajustar_altura_linha(current_row, descricao, altura_base)
            current_row += 1

            if progress_callback and total_linhas > 0:
                pct = int(((i + 1) / total_linhas) * 100)
                progress_callback(pct)

        return current_row, mapa_linhas_escritas

    def _processar_rodape(self, current_row: int, start_row: int) -> None:
        ultima_linha_dados = current_row - 1
        if not self.ws_out: return
        footer_config = self.perfil.get("rodape") if self.perfil else {"inicio": 26, "fim": 51}
        if not footer_config:
            return

        total_col = self._obter_col_saida("TOTAL")
        if not total_col:
            return

        footer_start = int(footer_config.get("inicio", 26))
        footer_end = int(footer_config.get("fim", 51))
        if footer_start < 1 or footer_end < footer_start:
            return

        total_letter = get_column_letter(total_col)
        max_row_planilha = max(self.ws_out.max_row, 200)
        self._limpar_area_total(current_row, max_row_planilha)

        target_start_row = current_row 
        self._copiar_bloco_excel(footer_start, footer_end, target_start_row)
        
        bdi_val = self._parse_num(self.info.get("bdi")) or Decimal("0")
        bdi_formula_value = self._decimal_to_formula_number(bdi_val)
        fator_desconto = Decimal("0.19") if abs(float(bdi_val) - 0.2882) < 0.001 else Decimal("0.0601")
        fator_desconto_formula = self._decimal_to_formula_number(fator_desconto)

        r1, r2, r3, r4, r5 = [target_start_row + i for i in range(5)]
        font_bold = Font(name="Arial", bold=True, size=10)

        target_sem_bdi = self._parse_num(self.info.get("orcafascio_total_sem_bdi"))
        target_bdi = self._parse_num(self.info.get("orcafascio_total_bdi"))

        subtotal_formula = f"=SUBTOTAL(9, {total_letter}{start_row}:{total_letter}{ultima_linha_dados})"
        subtotal_base = self._item_total_original_sum if self._item_total_original_sum > Decimal("0") else None
        if target_sem_bdi is not None and subtotal_base is not None:
            diff_sem_bdi = self._round_centavos(target_sem_bdi - subtotal_base)
            subtotal_formula = f"{subtotal_formula}{self._format_formula_delta(diff_sem_bdi)}"

        bdi_formula = f"=ROUND({total_letter}{r1}*{bdi_formula_value}, 2)"
        bdi_base = target_sem_bdi if target_sem_bdi is not None else subtotal_base
        if target_bdi is not None and bdi_base is not None:
            bdi_formula = self._get_exact_multiplier_formula(
                value_ref=f"{total_letter}{r1}",
                multiplier=bdi_val,
                target_total=target_bdi,
                value=bdi_base,
            )

        self._safe_write(r1, total_col, subtotal_formula)
        self._safe_write(r2, total_col, bdi_formula)
        self._safe_write(r3, total_col, f"={total_letter}{r1}+{total_letter}{r2}")
        self._safe_write(r4, total_col, f"=ROUNDDOWN({total_letter}{r3}*{fator_desconto_formula}, 2)")
        self._safe_write(r5, total_col, f"={total_letter}{r3}-{total_letter}{r4}")

        for r in [r1, r2, r3, r4, r5]:
            self.ws_out.row_dimensions[r].height = 30.0
            c = self.ws_out.cell(r, total_col)
            c.number_format = self.FMT_CONTABIL
            c.font = font_bold

    def _safe_write(self, row: int, col: int, value: Any, number_format: Optional[str] = None) -> None:
        if not self.ws_out: return
        try:
            cell = self.ws_out.cell(row, col)
            if isinstance(cell, MergedCell):
                for merged in list(self.ws_out.merged_cells.ranges):
                    if cell.coordinate in merged:
                        self.ws_out.unmerge_cells(str(merged))
                        break
                cell = self.ws_out.cell(row, col)
            
            if isinstance(cell, MergedCell):
                return

            cell.value = value
            if number_format: cell.number_format = number_format
        except Exception: 
            pass

    def _write_cell(self, coord: str, text: Any, bold: bool = True) -> None:
        if not self.ws_out: return
        try:
            cell = self.ws_out[coord]
            if isinstance(cell, MergedCell):
                for merged in self.ws_out.merged_cells.ranges:
                    if cell.coordinate in merged:
                        cell = self.ws_out.cell(merged.min_row, merged.min_col)
                        break
            
            cell.value = text
            if cell.font:
                current_font = copy(cell.font)
                new_font = Font(name=current_font.name, size=current_font.size, bold=bold, color=current_font.color)
                cell.font = new_font
        except Exception: 
            pass

    def _aplicar_estilo_hierarquico(self, row: int, nivel: str) -> None:
        if not self.ws_out: return
        paleta = {
            "N1": {"bg": "9BC2E6", "bold": True, "size": 11},
            "N2": {"bg": "BDD7EE", "bold": True, "size": 11},
            "N3": {"bg": "DDEBF7", "bold": True, "size": 11},
            "ITEM": {"bg": "FFFFFF", "bold": False, "size": 10}
        }
        estilo = paleta.get(nivel, paleta["ITEM"])
        fill = PatternFill("solid", fgColor=estilo["bg"])
        font = Font(name="Arial", bold=estilo["bold"], size=estilo["size"])
        side = Side(style="thin")
        border = Border(left=side, right=side, top=side, bottom=side)
        output_columns = sorted({col for col in self.mapa_saida.values() if col})
        if not output_columns:
            return

        descricao_col = self._obter_col_saida("DESCRICAO")
        total_col = self._obter_col_saida("TOTAL")

        for c in output_columns:
            try:
                cell = self.ws_out.cell(row, c)
                if isinstance(cell, MergedCell): continue
                cell.fill = fill
                cell.font = font
                cell.border = border
                h = "center"
                if c == descricao_col: h = "left"
                if c == total_col: h = "right"
                cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=(c == descricao_col))
            except: 
                pass

    def _ajustar_altura_linha(self, row: int, desc_txt: Any, altura_base: float) -> None:
        if not self.ws_out: return
        desc_txt = str(desc_txt)
        num_chars = len(desc_txt)
        linhas_estimadas = max(1, math.ceil(num_chars / 85))
        altura_final = altura_base if linhas_estimadas == 1 else max(altura_base, float(linhas_estimadas * 15))
        self.ws_out.row_dimensions[row].height = altura_final

    def _aplicar_precisao(self, valor: Optional[float], modo: str) -> Optional[float]:
        if valor is None: return None

        val_decimal = self._to_decimal(valor)
        if val_decimal is None:
            return None
            
        if modo == "TRUNC":
            return self._floor_centavos(val_decimal)
        elif modo == "ROUND": 
            return self._round_centavos(val_decimal)
        else: 
            return val_decimal

    def _limpar_area_total(self, row_inicio: int, row_fim: int) -> None:
        if not self.ws_out: return
        for merged in list(self.ws_out.merged_cells.ranges):
            if merged.max_row >= row_inicio:
                try: self.ws_out.unmerge_cells(str(merged))
                except: pass
        rows_to_delete = row_fim - row_inicio + 1
        if rows_to_delete > 0:
            self.ws_out.delete_rows(row_inicio, rows_to_delete)

    def _limpar_mesclagem_linha(self, row: int) -> None:
        if not self.ws_out: return
        for merged in list(self.ws_out.merged_cells.ranges):
            if row >= merged.min_row and row <= merged.max_row:
                try: self.ws_out.unmerge_cells(str(merged))
                except: pass

    def _copiar_bloco_excel(self, r_ini: int, r_fim: int, r_tgt_ini: int) -> None:
        if not self.ws_out or not self.ws_src: return
        offset = r_tgt_ini - r_ini
        for row in range(r_ini, r_fim + 1):
            tgt_row = row + offset
            self.ws_out.row_dimensions[tgt_row].height = self.ws_src.row_dimensions[row].height
            for col in range(1, self.ws_src.max_column + 1):
                cell_src = self.ws_src.cell(row, col)
                cell_tgt = self.ws_out.cell(tgt_row, col)
                if isinstance(cell_tgt, MergedCell):
                     for m in self.ws_out.merged_cells.ranges:
                         if cell_tgt.coordinate in m: 
                             self.ws_out.unmerge_cells(str(m))
                             break
                     cell_tgt = self.ws_out.cell(tgt_row, col)
                cell_tgt.value = cell_src.value
                if cell_src.has_style:
                    cell_tgt.font = copy(cell_src.font)
                    cell_tgt.border = copy(cell_src.border)
                    cell_tgt.fill = copy(cell_src.fill)
                    cell_tgt.number_format = cell_src.number_format
                    cell_tgt.alignment = copy(cell_src.alignment)
        for merged in self.ws_src.merged_cells.ranges:
            min_c, min_r, max_c, max_r = range_boundaries(str(merged))
            if min_r >= r_ini and max_r <= r_fim:
                new_min_r = min_r + offset
                new_max_r = max_r + offset
                coord_start = f"{get_column_letter(min_c)}{new_min_r}"
                coord_end = f"{get_column_letter(max_c)}{new_max_r}"
                try: self.ws_out.merge_cells(f"{coord_start}:{coord_end}")
                except ValueError: pass

    def _inserir_formulas_totais(self, mapa: List[Dict[str, Any]]) -> None:
        if not self.ws_out: return
        total_col = self._obter_col_saida("TOTAL")
        if not total_col:
            return

        total_letter = get_column_letter(total_col)
        try:
            total_rows = len(mapa)
            peso = {"N1": 1, "N2": 2, "N3": 3, "ITEM": 4}
            for i, atual in enumerate(mapa):
                if atual['nivel'] == "ITEM": continue
                row_pai = atual['row']
                nivel_pai = peso.get(atual['nivel'], 1)
                idx_fim = i
                for j in range(i + 1, total_rows):
                    prox = mapa[j]
                    if peso.get(prox['nivel'], 4) <= nivel_pai: break
                    idx_fim = j
                if idx_fim > i:
                    r_ini = mapa[i+1]['row']
                    r_fim = mapa[idx_fim]['row']
                    self._safe_write(row_pai, total_col, f"=SUBTOTAL(9, {total_letter}{r_ini}:{total_letter}{r_fim})", self.FMT_CONTABIL)
                    cell = self.ws_out.cell(row_pai, total_col)
                    if not isinstance(cell, MergedCell):
                        cell.font = Font(bold=True)
        except Exception as e: 
            Logger.error(f"Erro ao inserir totais sub-hierárquicos: {e}")

    def _calcular_total_formula(self, qtd: Optional[float], unit: Optional[float], calc_mode: str) -> Optional[float]:
        if qtd is None or unit is None:
            return None

        qtd_decimal = self._to_decimal(qtd)
        unit_decimal = self._to_decimal(unit)
        if qtd_decimal is None or unit_decimal is None:
            return None

        produto = qtd_decimal * unit_decimal
        if calc_mode == "TRUNC":
            return self._floor_centavos(produto)
        return self._round_centavos(produto)

    def _to_decimal(self, value: Any) -> Optional[Decimal]:
        if value is None:
            return None

        if isinstance(value, Decimal):
            return value

        try:
            if isinstance(value, float) and not math.isfinite(value):
                return None
        except TypeError:
            return None

        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _round_centavos(self, value: Any) -> Decimal:
        decimal_value = self._to_decimal(value)
        if decimal_value is None:
            return Decimal("0.00")
        return decimal_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _floor_centavos(self, value: Any) -> Decimal:
        decimal_value = self._to_decimal(value)
        if decimal_value is None:
            return Decimal("0.00")
        return decimal_value.quantize(Decimal("0.01"), rounding=ROUND_FLOOR)

    def _ceil_centavos(self, value: Any) -> Decimal:
        decimal_value = self._to_decimal(value)
        if decimal_value is None:
            return Decimal("0.00")
        return decimal_value.quantize(Decimal("0.01"), rounding=ROUND_CEILING)

    def _decimal_to_formula_number(self, value: Any) -> str:
        decimal_value = self._to_decimal(value) or Decimal("0")
        text = format(decimal_value.normalize(), "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text or "0"

    def _mesmo_centavo(self, left: Optional[float], right: Optional[float]) -> bool:
        if left is None or right is None:
            return False
        try:
            left_decimal = self._to_decimal(left)
            right_decimal = self._to_decimal(right)
            if left_decimal is None or right_decimal is None:
                return False
            return abs(left_decimal - right_decimal) <= Decimal("0.001")
        except (TypeError, ValueError, InvalidOperation):
            return False

    def _format_formula_delta(self, delta: Any) -> str:
        delta_decimal = self._to_decimal(delta) or Decimal("0")
        if abs(delta_decimal) <= Decimal("0.001"):
            return ""

        sign = "+" if delta_decimal > 0 else "-"
        return f"{sign}{self._decimal_to_formula_number(abs(delta_decimal))}"

    def _get_exact_formula(
        self,
        row: int,
        col_q: Optional[int],
        col_u: Optional[int],
        target_total: Optional[float],
        q_val: Optional[float],
        u_val: Optional[float],
    ) -> Optional[str]:
        if not col_q or not col_u:
            return None

        quant_ref = f"{get_column_letter(col_q)}{row}"
        unit_ref = f"{get_column_letter(col_u)}{row}"
        base_formula = f"=ROUND({quant_ref}*{unit_ref}, 2)"

        if target_total is None or q_val is None or u_val is None:
            return base_formula

        qtd_decimal = self._to_decimal(q_val)
        unit_decimal = self._to_decimal(u_val)
        target = self._to_decimal(target_total)
        if qtd_decimal is None or unit_decimal is None or target is None:
            return base_formula

        produto = qtd_decimal * unit_decimal
        candidates = [
            (self._round_centavos(produto), f"=ROUND({quant_ref}*{unit_ref}, 2)"),
            (self._floor_centavos(produto), f"=ROUNDDOWN({quant_ref}*{unit_ref}, 2)"),
            (self._ceil_centavos(produto), f"=ROUNDUP({quant_ref}*{unit_ref}, 2)"),
        ]

        for value, formula in candidates:
            if self._mesmo_centavo(value, target):
                return formula

        diff = self._round_centavos(target - produto)
        Logger.warning(
            f"Total original diverge dos arredondamentos nativos na linha {row}; "
            f"aplicando ajuste residual de {diff:+.2f} na fórmula."
        )
        return f"{base_formula}{self._format_formula_delta(diff)}"

    def _get_exact_multiplier_formula(
        self,
        value_ref: str,
        multiplier: Any,
        target_total: Optional[float],
        value: Optional[Any],
    ) -> str:
        multiplier_decimal = self._to_decimal(multiplier) or Decimal("0")
        multiplier_formula = self._decimal_to_formula_number(multiplier_decimal)
        base_formula = f"=ROUND({value_ref}*{multiplier_formula}, 2)"
        if target_total is None or value is None:
            return base_formula

        value_decimal = self._to_decimal(value)
        target = self._to_decimal(target_total)
        if value_decimal is None or target is None:
            return base_formula

        produto = value_decimal * multiplier_decimal
        candidates = [
            (self._round_centavos(produto), base_formula),
            (self._floor_centavos(produto), f"=ROUNDDOWN({value_ref}*{multiplier_formula}, 2)"),
            (self._ceil_centavos(produto), f"=ROUNDUP({value_ref}*{multiplier_formula}, 2)"),
        ]

        for projected, formula in candidates:
            if self._mesmo_centavo(projected, target):
                return formula

        diff = self._round_centavos(target - produto)
        Logger.warning(f"Total do BDI diverge da projeção; aplicando ajuste residual de {diff:+.2f}.")
        return f"{base_formula}{self._format_formula_delta(diff)}"

    def _resolver_total_item(
        self,
        qtd_final: Optional[float],
        unit_final: Optional[float],
        total_original: Optional[float],
        quant_ref: str,
        unit_ref: str,
        calc_mode: str,
    ) -> Any:
        if qtd_final is None or unit_final is None or not quant_ref or not unit_ref:
            return None

        produto = float(qtd_final) * float(unit_final)
        total_round = round(produto, 2)
        total_floor = math.floor(produto * 100) / 100.0

        if total_original is not None:
            if self._mesmo_centavo(total_round, total_original):
                return f"=ROUND({quant_ref}*{unit_ref}, 2)"

            if self._mesmo_centavo(total_floor, total_original):
                return f"=ROUNDDOWN({quant_ref}*{unit_ref}, 2)"

            delta_round = round(float(total_original) - total_round, 2)
            delta_floor = round(float(total_original) - total_floor, 2)
            use_floor = abs(delta_floor) < abs(delta_round)
            base_formula = "ROUNDDOWN" if use_floor else "ROUND"
            base_delta = delta_floor if use_floor else delta_round

            if abs(base_delta) <= 0.05:
                Logger.warning(
                    f"Total original diverge de ROUND/ROUNDDOWN na linha {quant_ref[1:] if quant_ref else '?'}; "
                    f"aplicando ajuste de {base_delta:+.2f} na fórmula."
                )
                return f"={base_formula}({quant_ref}*{unit_ref}, 2){self._format_formula_delta(base_delta)}"

            Logger.warning(
                "Total original não bate com ROUND/ROUNDDOWN e a diferença é maior que 5 centavos; "
                "mantendo fórmula padrão do modo de cálculo."
            )

        formula = "ROUNDDOWN" if calc_mode == "TRUNC" else "ROUND"
        return f"={formula}({quant_ref}*{unit_ref}, 2)"

    def _parse_num(self, val: Any) -> Optional[Decimal]:
        """Conversor financeiro seguro, sem round-half-even nem mutação por float."""
        if val is None:
            return None

        if isinstance(val, Decimal):
            return val

        try:
            import pandas as pd
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
            
        if isinstance(val, (int, float)):
            if isinstance(val, float) and not math.isfinite(val):
                return None
            return Decimal(str(val))

        try:
            s = str(val).upper().replace('R$', '').replace('\xa0', ' ').strip()
            if not s or s in {'NAN', 'NONE', 'NULL'}:
                return None

            match = re.search(r"-?\d[\d.\s]*,\d+|-?\d+(?:\.\d+)?", s)
            if not match:
                return None

            number = match.group(0).replace(' ', '')
            if ',' in number:
                number = number.replace('.', '').replace(',', '.')

            return Decimal(number)
        except (InvalidOperation, ValueError, TypeError):
            return None
