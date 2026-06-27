from fastapi import Depends, FastAPI, Request, UploadFile, File, Form, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from core.paths import get_resource_path, get_app_dir
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pathlib import Path
from contextlib import suppress
import json
import shutil
import uvicorn
import asyncio
import os
import sys
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from decimal import Decimal, InvalidOperation
import math
import re
import unicodedata
import uuid

import subprocess
from loguru import logger

from core.database import (
    BudgetHistory,
    SessionLocal,
    UserPreferences,
    get_or_create_preferences,
    get_session,
    init_database,
    record_budget_history,
    update_preferences,
)
from core.worker import worker_manager

CURRENT_VERSION = "0.5.0"

app = FastAPI(title="Planify v5.0 API")
init_database()

LOG_DIR = get_app_dir() / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
logger.add(
    LOG_DIR / "crash.log",
    rotation="5 MB",
    retention="30 days",
    backtrace=True,
    diagnose=True,
    enqueue=True,
    encoding="utf-8",
)


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.opt(exception=exc).error(
        "Unhandled backend exception on {method} {url}",
        method=request.method,
        url=str(request.url),
    )
    return JSONResponse(
        status_code=500,
        content={"error": "Erro interno inesperado. Consulte o crash.log para detalhes."},
    )

# Configuração de CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# WebSocket Connection Manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: list[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except Exception:
                pass

manager = ConnectionManager()
worker_event_pump_task: Optional[asyncio.Task] = None

# Pydantic Model for Budget Generation
class GerarOrcamentoRequest(BaseModel):
    table_data: List[Dict[str, Any]]
    mapping: Dict[str, str]
    side_data: Dict[str, Any]
    config_data: Dict[str, Any]
    caminho_sintetico: Optional[str] = None


class AutocompletePayload(BaseModel):
    sugestoes: Optional[Dict[str, List[str]]] = None
    autocomplete: Optional[Dict[str, List[str]]] = None


EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}
TEMPLATE_EXTENSIONS = {".xlsx", ".xlsm"}

CANONICAL_COLUMN_ALIASES = {
    "ITEM": ["ITEM", "ITEM ORCAMENTO", "ITEM ORÇAMENTO"],
    "CODIGO": ["CODIGO", "CÓDIGO", "COD", "COD.", "CODIGO SINAPI", "CÓDIGO SINAPI"],
    "BANCO": ["BANCO", "FONTE", "BASE", "REFERENCIA", "REFERÊNCIA"],
    "DESCRICAO": ["DESCRICAO", "DESCRIÇÃO", "DISCRIMINACAO", "DISCRIMINAÇÃO", "SERVICO", "SERVIÇO"],
    "UNID": ["UNID", "UND", "UNIDADE", "UN"],
    "QUANT": ["QUANT", "QUANT.", "QTD", "QTDE", "QUANTIDADE"],
    "UNIT": ["UNIT", "UNITARIO", "UNITÁRIO", "VALOR UNIT", "VALOR UNIT.", "PRECO UNITARIO", "PREÇO UNITÁRIO"],
    "TOTAL": ["TOTAL", "VALOR TOTAL", "PRECO TOTAL", "PREÇO TOTAL"],
}


def _normalize_column_name(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()
    return re.sub(r"\s+", " ", text)


def _has_known_headers(columns: List[Any]) -> bool:
    normalized = {_normalize_column_name(column) for column in columns}
    alias_sets = [
        {_normalize_column_name(alias) for alias in aliases}
        for aliases in CANONICAL_COLUMN_ALIASES.values()
    ]
    return sum(1 for aliases in alias_sets if normalized & aliases) >= 2


def _detect_header_row(raw_df: Any) -> Optional[int]:
    """Detecta a linha de cabeçalho quando o sintético vem com título antes da tabela."""
    alias_sets = [
        {_normalize_column_name(alias) for alias in aliases}
        for aliases in CANONICAL_COLUMN_ALIASES.values()
    ]

    for index, row in raw_df.head(40).iterrows():
        normalized_values = {
            _normalize_column_name(value)
            for value in row.tolist()
            if str(value).strip() and str(value).strip().lower() != "nan"
        }
        score = sum(1 for aliases in alias_sets if normalized_values & aliases)
        if score >= 2:
            return int(index)
    return None


def _make_unique_columns(columns: List[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []

    for index, column in enumerate(columns, start=1):
        name = str(column or "").strip()
        if not name or name.lower() == "nan":
            name = f"Coluna_{index}"

        count = seen.get(name, 0)
        seen[name] = count + 1
        result.append(name if count == 0 else f"{name}_{count + 1}")

    return result


def _select_best_html_table(tables: List[Any]) -> Any:
    """
    O Orçafascio exporta .xls que na verdade são HTML.
    Selecionamos a tabela com melhor chance de conter o sintético.
    """
    best_table = None
    best_score = -1

    for table in tables:
        if table is None or getattr(table, "empty", True):
            continue

        header_row = _detect_header_row(table)
        score = int(table.shape[0]) + int(table.shape[1])
        if header_row is not None:
            score += 1000

        if score > best_score:
            best_table = table
            best_score = score

    if best_table is None:
        raise ValueError("Nenhuma tabela válida foi encontrada no HTML exportado.")

    return best_table


def _read_html_tables(pd: Any, file_path: Path) -> List[Any]:
    last_error: Optional[Exception] = None

    for encoding in ("utf-8", "cp1252", "latin1"):
        try:
            tables = pd.read_html(
                file_path,
                header=None,
                encoding=encoding,
                keep_default_na=False,
                thousands=None,
                converters={index: str for index in range(200)},
            )
            if tables:
                return tables
        except Exception as exc:
            last_error = exc

    if last_error:
        raise last_error

    return []


def _promote_detected_header(raw_df: Any) -> Any:
    header_row = _detect_header_row(raw_df)
    if header_row is None:
        raw_df = raw_df.copy()
        raw_df.columns = _make_unique_columns(list(raw_df.columns))
        return raw_df

    df = raw_df.iloc[header_row + 1:].copy()
    df.columns = _make_unique_columns(raw_df.iloc[header_row].tolist())
    return df


def _read_excel_dataframe(pd: Any, file_path: Path, engine: Optional[str] = None) -> Any:
    # Lemos como texto para preservar exatamente o que o usuário enxerga no
    # sintético. A conversão numérica oficial fica centralizada no engine com Decimal.
    read_kwargs: Dict[str, Any] = {
        "dtype": str,
        "keep_default_na": False,
        "na_filter": False,
    }
    if engine:
        read_kwargs["engine"] = engine

    df = pd.read_excel(file_path, **read_kwargs)

    if not _has_known_headers(list(df.columns)):
        raw_df = pd.read_excel(file_path, header=None, **read_kwargs)
        header_row = _detect_header_row(raw_df)
        if header_row is not None:
            df = pd.read_excel(file_path, header=header_row, **read_kwargs)

    return df


def _read_orcafascio_dataframe(pd: Any, file_path: Path) -> tuple[Any, str]:
    """
    Lê o sintético sem depender do Excel nativo.
    Fluxo:
    1. pandas/openpyxl para arquivos Excel íntegros;
    2. calamine para xlsx com XML visual corrompido;
    3. read_html para .xls do Orçafascio que na verdade é HTML.
    """
    try:
        return _read_excel_dataframe(pd, file_path), "excel"
    except Exception as excel_error:
        calamine_error: Optional[Exception] = None

        try:
            return _read_excel_dataframe(pd, file_path, engine="calamine"), "excel_calamine"
        except Exception as exc:
            calamine_error = exc
            logger.warning(
                "Leitura com Calamine falhou para {file}: {error}",
                file=str(file_path),
                error=exc,
            )

        try:
            tables = _read_html_tables(pd, file_path)
        except ImportError:
            raise
        except Exception as html_error:
            raise ValueError(
                "Falha ao ler o arquivo. O Excel normal falhou, o motor Calamine "
                "não conseguiu extrair os dados e o arquivo também não pôde ser "
                "interpretado como HTML do Orçafascio. "
                f"Excel: {excel_error}; Calamine: {calamine_error}; HTML: {html_error}"
            ) from html_error

        if not tables:
            raise ValueError(
                "O arquivo parece ser HTML, mas nenhuma tabela foi encontrada no export do Orçafascio."
            )

        raw_df = _select_best_html_table(tables)
        return _promote_detected_header(raw_df), "html"


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True

    if isinstance(value, float) and not math.isfinite(value):
        return True

    text = str(value).strip()
    return text == "" or text.lower() in {"nan", "none", "null"}


def _parse_orcafascio_number(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return str(value)

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return None

    text = text.replace("R$", "").replace("\xa0", " ")
    match = re.search(r"-?\d[\d.\s]*,\d{1,4}|-?\d+(?:\.\d+)?", text)
    if not match:
        return None

    number = match.group(0).replace(" ", "")
    if "," in number:
        number = number.replace(".", "").replace(",", ".")

    try:
        Decimal(number)
        return number
    except (InvalidOperation, ValueError):
        return None


def _extract_footer_value_from_row(row_values: List[Any], label_index: int) -> Optional[str]:
    candidates = row_values[label_index + 1:] or row_values
    for value in reversed(candidates):
        parsed = _parse_orcafascio_number(value)
        if parsed is not None:
            return parsed
    return None


def _extract_footer_totals_and_cut(df: Any) -> tuple[Any, Dict[str, str]]:
    """
    O sintético do Orçafascio traz um rodapé próprio depois de "Total sem BDI".
    Essas linhas não são itens orçamentários e corrompem o preview/geração.
    Antes do corte, os totais originais são resgatados para fechar o Excel gerado
    com fórmulas compensadas, sem colar valores estáticos.
    """
    footer_info: Dict[str, str] = {}
    cut_position: Optional[int] = None

    for position, (_, row) in enumerate(df.iterrows()):
        row_values = row.tolist() if hasattr(row, "tolist") else list(row)
        normalized_values = [_normalize_column_name(value) for value in row_values]

        for index, normalized in enumerate(normalized_values):
            if "TOTAL SEM BDI" in normalized:
                value = _extract_footer_value_from_row(row_values, index)
                if value is not None:
                    footer_info["orcafascio_total_sem_bdi"] = value
                cut_position = position if cut_position is None else min(cut_position, position)
                break

        for index, normalized in enumerate(normalized_values):
            if "TOTAL DO BDI" in normalized or normalized == "TOTAL BDI":
                value = _extract_footer_value_from_row(row_values, index)
                if value is not None:
                    footer_info["orcafascio_total_bdi"] = value
                break

    if cut_position is not None:
        return df.iloc[:cut_position].copy(), footer_info

    return df, footer_info


def _item_hierarchy_level(item_value: Any) -> str:
    if isinstance(item_value, (int, float)) and math.isfinite(float(item_value)):
        text = str(int(item_value)) if float(item_value).is_integer() else str(item_value)
    else:
        text = str(item_value or "").strip()

    if not text:
        return "N1"

    # Orçafascio pode entregar "1.", "1.1." ou valores normalizados como "1.1".
    parts = [part for part in text.rstrip(".").split(".") if part.strip()]
    if len(parts) <= 1:
        return "N1"
    if len(parts) == 2:
        return "N2"
    return "N3"


def _apply_forced_hierarchy(rows: List[Dict[str, Any]], mapping: Dict[str, str]) -> List[Dict[str, Any]]:
    """
    Marca linhas-título para que o motor de Excel aplique as cores N1/N2/N3.
    Uma etapa não possui código, banco, unidade e valor unitário.
    """
    item_col = mapping.get("ITEM") or ""
    title_probe_columns = [
        mapping.get("CODIGO") or "",
        mapping.get("BANCO") or "",
        mapping.get("UNID") or "",
        mapping.get("UNIT") or "",
    ]

    for row in rows:
        is_title_row = bool(item_col and str(row.get(item_col, "")).strip())
        is_title_row = is_title_row and all(
            not column or _is_blank_cell(row.get(column))
            for column in title_probe_columns
        )
        row["_NIVEL_FORCADO"] = _item_hierarchy_level(row.get(item_col)) if is_title_row else "ITEM"

    return rows


def _json_safe_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, float) and not math.isfinite(value):
        return ""

    if hasattr(value, "item"):
        try:
            value = value.item()
        except (ValueError, TypeError):
            pass

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (ValueError, TypeError):
            return str(value)

    if isinstance(value, dict):
        return {str(key): _json_safe_value(val) for key, val in value.items()}

    if isinstance(value, (list, tuple)):
        return [_json_safe_value(item) for item in value]

    return value


def _infer_column_mapping(rows: List[Dict[str, Any]], provided_mapping: Optional[Dict[str, str]] = None) -> Dict[str, str]:
    provided_mapping = provided_mapping or {}
    columns: List[str] = []

    for row in rows:
        if row:
            columns = [str(column) for column in row.keys()]
            break

    normalized_to_original = {_normalize_column_name(column): column for column in columns}
    inferred: Dict[str, str] = {}

    for canonical, aliases in CANONICAL_COLUMN_ALIASES.items():
        requested = provided_mapping.get(canonical)
        if requested and requested in columns:
            inferred[canonical] = requested
            continue

        for alias in aliases:
            original = normalized_to_original.get(_normalize_column_name(alias))
            if original:
                inferred[canonical] = original
                break

        inferred.setdefault(canonical, canonical if canonical in columns else "")

    return inferred


def _normalize_generation_config(config_data: Dict[str, Any]) -> Dict[str, Any]:
    config = dict(config_data or {})

    try:
        raw_bdi = str(config.get("bdi", "0.2882")).replace(",", ".").strip()
        bdi = float(raw_bdi) if raw_bdi else 0.2882
        config["bdi"] = bdi / 100 if bdi > 1 else bdi
    except (TypeError, ValueError):
        config["bdi"] = 0.2882

    method_map = {
        "cortar": "TRUNC",
        "trunc": "TRUNC",
        "arredondar": "ROUND",
        "round": "ROUND",
        "exato": "EXACT",
        "exact": "EXACT",
    }
    raw_method = str(config.get("metodo_calculo", config.get("calc_mode", "exato"))).lower()
    config["calc_mode"] = method_map.get(raw_method, "EXACT")

    try:
        config["altura_linha"] = float(str(config.get("altura_linha", 24.75)).replace(",", "."))
    except (TypeError, ValueError):
        config["altura_linha"] = 24.75

    return config


def _extract_template_preview(file_path: Path) -> Dict[str, Any]:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        max_column = min(int(worksheet.max_column or 0), 80)
        max_row = min(int(worksheet.max_row or 0), 20)

        columns = [
            {"letter": get_column_letter(index), "index": index}
            for index in range(1, max_column + 1)
        ]
        rows = []

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            max_col=max_column,
            values_only=True,
        ):
            rows.append([_json_safe_value(value) for value in row])

        return {
            "sheet_name": worksheet.title,
            "columns": columns,
            "rows": rows,
            "max_column": int(worksheet.max_column or 0),
            "max_row": int(worksheet.max_row or 0),
        }
    finally:
        workbook.close()


def _budget_output_suffix(profile: Dict[str, Any]) -> str:
    template_path = str(profile.get("caminho_template") or profile.get("filepath") or "")
    return ".xlsm" if template_path.lower().endswith(".xlsm") else ".xlsx"


def _safe_budget_filename(side_data: Dict[str, Any], suffix: str) -> str:
    raw_name = str(side_data.get("nome_arquivo") or side_data.get("processo") or "Orcamento").strip()
    stem = Path(raw_name).stem or "Orcamento"
    stem = re.sub(r'[<>:"/\\|?*]', "", stem).strip(" .") or "Orcamento"
    return f"{stem}{suffix}"


def _next_available_output_path(default_path: Path) -> Path:
    if not default_path.exists():
        return default_path

    for attempt in range(1, 100):
        candidate = default_path.with_name(f"{default_path.stem}_v{attempt}{default_path.suffix}")
        if not candidate.exists():
            return candidate

    return default_path.with_name(f"{default_path.stem}_{uuid.uuid4().hex[:8]}{default_path.suffix}")


def _normalize_selected_save_path(raw_path: Any, suffix: str) -> Path:
    if isinstance(raw_path, (list, tuple)):
        raw_path = raw_path[0] if raw_path else ""

    selected = Path(str(raw_path or "").strip()).expanduser()
    if not str(selected):
        raise ValueError("Nenhum caminho foi selecionado.")

    if selected.suffix.lower() not in {".xlsx", ".xlsm"}:
        selected = selected.with_suffix(suffix)
    elif suffix == ".xlsm" and selected.suffix.lower() != ".xlsm":
        selected = selected.with_suffix(".xlsm")

    selected.parent.mkdir(parents=True, exist_ok=True)
    return selected


def _select_save_path_for_budget(side_data: Dict[str, Any], profile: Dict[str, Any]) -> str:
    suffix = _budget_output_suffix(profile)
    output_dir = get_app_dir() / "Output"
    output_dir.mkdir(parents=True, exist_ok=True)

    fallback_path = _next_available_output_path(output_dir / _safe_budget_filename(side_data, suffix))

    try:
        import webview

        if not getattr(webview, "windows", None):
            raise RuntimeError("Janela PyWebView não está disponível para abrir o diálogo Salvar Como.")

        file_dialog = getattr(getattr(webview, "FileDialog", None), "SAVE", None)
        if file_dialog is None:
            file_dialog = getattr(webview, "SAVE_DIALOG")

        selected_path = webview.windows[0].create_file_dialog(
            file_dialog,
            directory=str(output_dir),
            save_filename=fallback_path.name,
            file_types=("Excel (*.xlsx;*.xlsm)",),
        )
        if not selected_path:
            raise HTTPException(status_code=400, detail="Geração cancelada pelo usuário.")

        return str(_normalize_selected_save_path(selected_path, suffix))
    except HTTPException:
        raise
    except Exception as exc:
        logger.warning(
            "Falha ao abrir diálogo nativo Salvar Como; usando fallback em Output: {error}",
            error=exc,
        )
        return str(fallback_path)


def _parse_total_value(value: Any) -> float:
    if value in (None, ""):
        return 0.0

    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)

    text = str(value).replace("R$", "").strip()
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def _record_task_history(task_id: str, result: Dict[str, Any]) -> None:
    payload = worker_manager.get_task_payload(task_id)
    if not payload:
        logger.warning("Payload da tarefa {task_id} não encontrado para histórico.", task_id=task_id)
        return

    side_data = dict(payload.get("side_data") or {})
    config_data = dict(payload.get("config_data") or {})
    profile = dict(payload.get("profile") or {})

    with SessionLocal() as session:
        record_budget_history(
            session,
            project_name=side_data.get("descricao_header") or side_data.get("nome_arquivo") or "",
            sector=side_data.get("setor") or "",
            total_value=_parse_total_value(side_data.get("valor_simulado")),
            excel_path=result.get("caminho_excel") or "",
            profile_id=profile.get("id"),
        )
        update_preferences(
            session,
            last_bdi=float(config_data.get("bdi", 0.2882)),
            last_profile_id=profile.get("id"),
        )


async def _handle_worker_event(event: Dict[str, Any]) -> None:
    event_type = event.get("type")
    task_id = event.get("task_id")

    if event_type == "task" and task_id:
        result = dict(event.get("result") or {})
        status = event.get("status")

        if status == "completed":
            try:
                _record_task_history(task_id, result)
            except Exception as exc:
                logger.opt(exception=exc).error("Falha ao gravar histórico da tarefa {task_id}", task_id=task_id)

            await manager.broadcast({
                "type": "log",
                "task_id": task_id,
                "message": "Orçamento finalizado com sucesso.",
                "level": "SUCCESS",
            })
        elif status == "failed":
            await manager.broadcast({
                "type": "log",
                "task_id": task_id,
                "message": event.get("error") or "Falha ao gerar orçamento.",
                "level": "ERROR",
            })

        await manager.broadcast(event)
        worker_manager.forget_task(task_id)
        return

    await manager.broadcast(event)


async def _worker_event_pump() -> None:
    while True:
        for event in worker_manager.drain_events():
            try:
                await _handle_worker_event(event)
            except Exception as exc:
                logger.opt(exception=exc).error("Falha ao processar evento do worker: {event}", event=event)
        await asyncio.sleep(0.15)


@app.on_event("startup")
async def start_worker_event_pump() -> None:
    global worker_event_pump_task
    if worker_event_pump_task is None or worker_event_pump_task.done():
        worker_event_pump_task = asyncio.create_task(_worker_event_pump())


@app.on_event("shutdown")
async def stop_worker_event_pump() -> None:
    global worker_event_pump_task
    if worker_event_pump_task:
        worker_event_pump_task.cancel()
        with suppress(asyncio.CancelledError):
            await worker_event_pump_task
        worker_event_pump_task = None

    worker_manager.shutdown()


@app.websocket("/ws/logs")
async def websocket_logs(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # We don't expect messages from the client, but we need to keep connection open
            data = await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)

class ExtrairTextoRequest(BaseModel):
    texto: str

@app.post("/api/extrair-texto")
async def extrair_texto(request: ExtrairTextoRequest):
    """
    Recebe um texto e extrai os dados normalizados.
    """
    try:
        from utils.smart_parser import SmartParser
        dados = SmartParser.parse_whatsapp_text(request.texto)
        return dados
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao extrair texto: {str(e)}")


@app.get("/api/ler-excel")
async def ler_excel(caminho: str):
    """
    Lê o arquivo Excel salvo e retorna dados seguros para o React.
    """
    if not caminho or not caminho.strip():
        raise HTTPException(status_code=400, detail="Caminho do arquivo não fornecido.")

    try:
        file_path = Path(caminho).expanduser().resolve(strict=True)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Arquivo não encontrado no servidor.")

    if not file_path.is_file():
        raise HTTPException(status_code=400, detail="O caminho informado não aponta para um arquivo.")

    if file_path.suffix.lower() not in EXCEL_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Extensão de arquivo não suportada.")

    try:
        # Lazy loading: pandas só é importado quando o usuário realmente lê Excel.
        import pandas as pd

        df, origem = _read_orcafascio_dataframe(pd, file_path)

        df = df.dropna(how="all")
        df = df.fillna("")
        df.columns = _make_unique_columns([str(column).strip() for column in df.columns])
        df, footer_info = _extract_footer_totals_and_cut(df)

        records = [
            {str(key): _json_safe_value(value) for key, value in row.items()}
            for row in df.to_dict(orient="records")
        ]
        suggested_mapping = _infer_column_mapping(records)
        records = _apply_forced_hierarchy(records, suggested_mapping)

        return {
            "status": "success",
            "dados": records,
            "colunas": [str(column) for column in df.columns],
            "linhas": len(records),
            "origem": origem,
            "mapeamento_sugerido": suggested_mapping,
            "info": footer_info,
        }
    except ImportError:
        raise HTTPException(status_code=500, detail="Dependência pandas/lxml/html5lib não instalada.")
    except ValueError as e:
        raise HTTPException(status_code=422, detail=f"Não foi possível interpretar o Excel: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler Excel: {str(e)}")

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """
    Recebe um arquivo Excel e o grava temporariamente.
    """
    if not file.filename or Path(file.filename).suffix.lower() not in EXCEL_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Extensão de arquivo não suportada. Use .xlsx, .xls ou .xlsm.")

    try:
        temp_dir = get_app_dir() / "Output" / "temp"
        temp_dir.mkdir(parents=True, exist_ok=True)

        # Sanitização contra path traversal e colisões entre uploads com o mesmo nome.
        original_filename = Path(file.filename).name
        safe_filename = re.sub(r"[^A-Za-z0-9._ -]", "_", original_filename).strip(" .")
        safe_filename = safe_filename or f"planilha{Path(original_filename).suffix.lower()}"
        file_path = temp_dir / f"{uuid.uuid4().hex}_{safe_filename}"
        
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        return {
            "status": "sucesso",
            "mensagem": f"Arquivo {original_filename} salvo com sucesso.",
            "caminho": str(file_path),
            "nome_arquivo": original_filename,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar o arquivo: {str(e)}")


@app.get("/api/autocomplete")
async def listar_autocomplete():
    """
    Retorna listas de preenchimento rápido sem acoplar o frontend ao JSON legado.
    """
    try:
        from utils.autocomplete_manager import AutocompleteManager

        autocomplete_manager = AutocompleteManager()
        suggestions = autocomplete_manager.get_all()
        return {
            "status": "success",
            "sugestoes": suggestions,
            "autocomplete": suggestions,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao carregar autocomplete: {str(e)}")


@app.put("/api/autocomplete")
async def atualizar_autocomplete(payload: AutocompletePayload):
    """
    Substitui permanentemente o banco de sugestões usado pelos comboboxes.
    Aceita {sugestoes: {...}} ou o objeto direto em {autocomplete: {...}}.
    """
    try:
        from utils.autocomplete_manager import AutocompleteManager

        raw_suggestions = payload.sugestoes if payload.sugestoes is not None else payload.autocomplete
        if raw_suggestions is None:
            raise HTTPException(status_code=400, detail="Envie o objeto de sugestões em 'sugestoes' ou 'autocomplete'.")

        autocomplete_manager = AutocompleteManager()
        suggestions = autocomplete_manager.replace_all(raw_suggestions)
        return {
            "status": "success",
            "sugestoes": suggestions,
            "autocomplete": suggestions,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar autocomplete: {str(e)}")


@app.post("/api/upload-template-preview")
async def upload_template_preview(file: UploadFile = File(...)):
    """
    Recebe um template Excel, salva temporariamente e devolve uma prévia
    das primeiras linhas/colunas para guiar o mapeamento visual.
    """
    if not file.filename or Path(file.filename).suffix.lower() not in TEMPLATE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Extensão de template não suportada. Use .xlsx ou .xlsm.")

    temp_path = None

    try:
        temp_dir = get_app_dir() / "Output" / "temp" / "templates"
        temp_dir.mkdir(parents=True, exist_ok=True)

        original_filename = Path(file.filename).name
        safe_name = re.sub(r"[^A-Za-z0-9._ -]", "_", original_filename).strip(" .") or "template.xlsx"
        temp_path = temp_dir / f"{uuid.uuid4().hex}_{safe_name}"

        with temp_path.open("wb") as output:
            shutil.copyfileobj(file.file, output)

        preview = await asyncio.to_thread(_extract_template_preview, temp_path)
        return {
            "status": "success",
            "caminho_template": str(temp_path),
            "nome_arquivo": original_filename,
            "preview": preview,
        }
    except HTTPException:
        if temp_path and temp_path.exists():
            with suppress(Exception):
                temp_path.unlink()
        raise
    except Exception as e:
        if temp_path and temp_path.exists():
            with suppress(Exception):
                temp_path.unlink()
        raise HTTPException(status_code=422, detail=f"Não foi possível ler o template Excel: {str(e)}")


@app.get("/api/perfis")
async def listar_perfis(session: Session = Depends(get_session)):
    """
    Lista os perfis de empresa disponíveis para o Universal Mapper.
    """
    try:
        from utils.template_manager import TemplateManager

        template_manager = TemplateManager(session)
        preferences = get_or_create_preferences(session)
        return {
            "status": "success",
            "perfis": template_manager.list_profiles(),
            "preferencias": {
                "last_bdi": preferences.last_bdi,
                "last_profile_id": preferences.last_profile_id,
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao listar perfis: {str(e)}")


@app.post("/api/perfis")
async def criar_perfil(
    request: Request,
    file: Optional[UploadFile] = File(None),
    perfil: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    """
    Recebe um perfil novo de duas formas:
    - multipart/form-data legado: file + perfil JSON;
    - application/json: nome_empresa, caminho_template, linha_inicio e mapa_colunas.
    """
    try:
        from utils.template_manager import TemplateManager

        content_type = request.headers.get("content-type", "").lower()
        if content_type.startswith("application/json"):
            payload = await request.json()
        else:
            if not perfil:
                raise HTTPException(status_code=400, detail="Envie os dados do perfil.")
            try:
                payload = json.loads(perfil)
            except json.JSONDecodeError:
                raise HTTPException(status_code=400, detail="O campo perfil deve conter um JSON válido.")

        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Payload do perfil inválido.")

        template_manager = TemplateManager(session)
        if file is not None:
            if not file.filename or Path(file.filename).suffix.lower() not in TEMPLATE_EXTENSIONS:
                raise HTTPException(status_code=400, detail="Extensão de template não suportada. Use .xlsx ou .xlsm.")
            ok, message, saved_profile = template_manager.add_profile(file.file, file.filename, payload)
        else:
            template_path = payload.get("caminho_template") or payload.get("filepath")
            ok, message, saved_profile = template_manager.add_profile_from_path(template_path, payload)

        if not ok:
            raise HTTPException(status_code=400, detail=message)

        return {
            "status": "success",
            "mensagem": message,
            "perfil": saved_profile,
            "perfis": template_manager.list_profiles(),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao salvar perfil: {str(e)}")


@app.delete("/api/perfis/{perfil_id}")
async def excluir_perfil(perfil_id: str, session: Session = Depends(get_session)):
    """
    Remove um perfil e o template associado da pasta gravável do usuário.
    """
    try:
        from utils.template_manager import TemplateManager

        template_manager = TemplateManager(session)
        if not template_manager.remove_profile(perfil_id):
            raise HTTPException(status_code=404, detail="Perfil não encontrado.")

        return {"status": "success", "perfis": template_manager.list_profiles()}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao remover perfil: {str(e)}")


@app.get("/api/version")
async def get_version():
    return {"version": CURRENT_VERSION}


@app.get("/api/health")
async def get_health():
    return {
        "status": "ok",
        "version": CURRENT_VERSION,
        "app_dir": str(get_app_dir()),
        "worker": "available",
    }


@app.delete("/api/historico")
async def limpar_historico(session: Session = Depends(get_session)):
    try:
        deleted = session.query(BudgetHistory).delete(synchronize_session=False)
        session.commit()
        return {"status": "success", "deleted": deleted}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao limpar histórico: {str(e)}")


@app.post("/api/preferencias/reset")
async def resetar_preferencias(session: Session = Depends(get_session)):
    try:
        session.query(UserPreferences).delete(synchronize_session=False)
        preferences = UserPreferences(id=1, last_bdi=0.0, last_profile_id=None)
        session.add(preferences)
        session.commit()
        return {"status": "success", "preferencias": {"last_bdi": 0.0, "last_profile_id": None}}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao resetar preferências: {str(e)}")


class UpdateRequest(BaseModel):
    download_url: str

@app.post("/api/update")
async def trigger_update(request: UpdateRequest):
    """
    Inicia o processo de auto-update baixando o novo binário
    e criando um script bat para substituir o .exe atual.
    """
    try:
        import urllib.request
        from urllib.parse import urlparse

        parsed_url = urlparse(request.download_url)
        if parsed_url.scheme not in {"http", "https"}:
            raise HTTPException(status_code=400, detail="URL de atualização inválida.")

        if getattr(sys, "frozen", False):
            executable_path = Path(sys.executable).resolve()
        else:
            executable_path = Path.cwd() / "PlanifyV5.exe"

        install_dir = executable_path.parent
        new_executable = install_dir / "PlanifyV5_novo.exe"
        updater_script = install_dir / "updater.bat"

        asyncio.run_coroutine_threadsafe(
            manager.broadcast({"type": "log", "message": "Baixando atualização...", "level": "INFO"}),
            asyncio.get_running_loop()
        )

        await asyncio.to_thread(urllib.request.urlretrieve, request.download_url, str(new_executable))

        bat_content = f"""@echo off
set "TARGET={executable_path}"
set "NEW_EXE={new_executable}"
set "INSTALL_DIR={install_dir}"

timeout /t 3 /nobreak > NUL
:wait_delete
del /f /q "%TARGET%" > NUL 2>&1
if exist "%TARGET%" (
  timeout /t 2 /nobreak > NUL
  goto wait_delete
)

move /y "%NEW_EXE%" "%TARGET%" > NUL
start "" "%TARGET%"
del "%~f0"
"""
        with open(updater_script, "w", encoding="utf-8") as f:
            f.write(bat_content)

        creationflags = subprocess.CREATE_NEW_CONSOLE if os.name == "nt" else 0
        subprocess.Popen(
            [str(updater_script)],
            shell=True,
            cwd=str(install_dir),
            creationflags=creationflags,
        )
        os._exit(0)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/abrir-pasta")
async def abrir_pasta(caminho: str):
    """
    Abre o Windows Explorer (ou equivalente do SO) selecionando o arquivo
    """
    try:
        if os.name == 'nt':
            # On Windows, selecting the file in explorer
            import subprocess
            subprocess.run(['explorer', '/select,', os.path.normpath(caminho)])
        elif os.name == 'posix':
            import subprocess
            # Mac or Linux
            if sys.platform == 'darwin':
                subprocess.run(['open', '-R', caminho])
            else:
                subprocess.run(['xdg-open', os.path.dirname(caminho)])
        return {"status": "sucesso"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/gerar-orcamento")
async def gerar_orcamento(request: GerarOrcamentoRequest, session: Session = Depends(get_session)):
    """
    Agenda a geração do orçamento em um processo isolado e retorna imediatamente.
    """
    try:
        if not request.table_data:
            await manager.broadcast({
                "type": "log",
                "message": "Nenhuma linha do Excel foi carregada. Faça o upload e aguarde o preview antes de gerar.",
                "level": "ERROR",
            })
            raise HTTPException(
                status_code=400,
                detail="Nenhuma linha do Excel foi carregada. Faça o upload e aguarde o preview antes de gerar.",
            )

        table_data = request.table_data
        mapping = _infer_column_mapping(table_data, request.mapping)
        side_data = dict(request.side_data or {})
        config_data = _normalize_generation_config(request.config_data)
        profile_id = str(config_data.get("perfil_id") or config_data.get("modelo") or "").strip()

        from utils.template_manager import TemplateManager

        template_manager = TemplateManager(session)
        profile = template_manager.get_profile(profile_id) if profile_id else None
        if not profile:
            available_profiles = template_manager.list_profiles()
            if len(available_profiles) == 1:
                profile = available_profiles[0]
            else:
                raise HTTPException(
                    status_code=400,
                    detail="Selecione um perfil de empresa válido antes de gerar o orçamento.",
                )

        config_data["perfil_id"] = profile["id"]
        config_data["modelo"] = profile["id"]
        config_data["save_path"] = _select_save_path_for_budget(side_data, profile)

        payload = {
            "table_data": table_data,
            "mapping": mapping,
            "side_data": side_data,
            "config_data": config_data,
            "profile": profile,
            "caminho_sintetico": request.caminho_sintetico,
        }
        task_id = worker_manager.submit_budget_task(payload)

        await manager.broadcast({
            "type": "task",
            "task_id": task_id,
            "status": "processing",
            "message": "Geração enviada para o worker isolado.",
        })
        await manager.broadcast({
            "type": "log",
            "task_id": task_id,
            "message": "Geração enviada para o worker isolado.",
            "level": "INFO",
        })

        return JSONResponse(
            status_code=202,
            content={
                "status": "Processing",
                "task_id": task_id,
                "message": "Geração enviada para o worker isolado.",
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        await manager.broadcast({"type": "log", "message": f"Exceção crítica: {str(e)}", "level": "ERROR"})
        raise HTTPException(status_code=500, detail=str(e))


# Serve Static Frontend Files
frontend_dist = get_resource_path("frontend/dist")

# We mount the static directory at /assets (or wherever Vite puts its assets)
# but for a SPA, we need a catch-all for index.html.

if frontend_dist.exists() and frontend_dist.is_dir():
    app.mount("/assets", StaticFiles(directory=frontend_dist / "assets"), name="assets")

    # Catch-all for SPA React Routing
    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        # Prevent accessing paths outside the frontend directory (e.g. if somehow asked)
        # Check if the requested file exists (like favicon.ico)
        requested_file = frontend_dist / full_path
        if requested_file.is_file():
            return FileResponse(requested_file)

        # Fallback to index.html for React Router
        return FileResponse(frontend_dist / "index.html")

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
